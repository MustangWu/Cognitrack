#!/usr/bin/env python3
"""
xlsx_to_sql.py — Generate SQL migration scripts from xlsx files.

Usage:
    python3 db/xlsx_to_sql.py <file.xlsx> [options]

Options:
    --sheet <name|index>   Sheet to process (default: all sheets)
    --table <name>         Override table name (default: sheet name, snake_cased)
    --dialect <pg|mysql>   SQL dialect (default: pg)
    --out <dir>            Output directory (default: same dir as xlsx file)
    --header <row>         Row number for header (1-based, default: 1)
    --no-drop              Skip DROP TABLE IF EXISTS

Examples:
    python3 db/xlsx_to_sql.py data/aihw_data.xlsx
    python3 db/xlsx_to_sql.py data/aihw_data.xlsx --sheet "Prevalence" --dialect mysql
    python3 db/xlsx_to_sql.py data/aihw_data.xlsx --out db/
"""

import argparse
import os
import re
import sys
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def to_snake_case(name: str) -> str:
    """Convert any string to a safe snake_case SQL identifier."""
    name = str(name).strip()
    name = re.sub(r"[^\w\s]", "", name)          # strip special chars
    name = re.sub(r"\s+", "_", name)              # spaces → underscore
    name = re.sub(r"([A-Z]+)([A-Z][a-z])", r"\1_\2", name)
    name = re.sub(r"([a-z\d])([A-Z])", r"\1_\2", name)
    name = name.lower().strip("_")
    name = re.sub(r"_+", "_", name)
    if name and name[0].isdigit():
        name = "col_" + name
    return name or "col"


def infer_sql_type(values: list, dialect: str) -> str:
    """
    Infer the tightest SQL column type from a list of non-null cell values.
    Priority: BOOLEAN → SMALLINT → INTEGER → BIGINT → NUMERIC → TEXT
    """
    bool_vals = {True, False, "true", "false", "yes", "no", "1", "0", 1, 0}

    all_bool = all(
        str(v).strip().lower() in {"true", "false", "yes", "no"} or v in {True, False}
        for v in values
    )
    if all_bool:
        return "BOOLEAN" if dialect == "pg" else "TINYINT(1)"

    # Try integer tiers
    int_ok = True
    max_abs = 0
    for v in values:
        if isinstance(v, bool):
            int_ok = False
            break
        try:
            i = int(float(str(v)))
            if float(str(v)) != i:
                int_ok = False
                break
            max_abs = max(max_abs, abs(i))
        except (ValueError, TypeError):
            int_ok = False
            break

    if int_ok:
        if max_abs <= 32767:
            return "SMALLINT"
        if max_abs <= 2147483647:
            return "INTEGER"
        return "BIGINT"

    # Try numeric/decimal
    num_ok = True
    max_scale = 0
    max_prec = 0
    for v in values:
        try:
            d = Decimal(str(v))
            sign, digits, exp = d.as_tuple()
            scale = max(0, -exp)
            prec = len(digits)
            max_scale = max(max_scale, scale)
            max_prec = max(max_prec, prec)
        except (InvalidOperation, TypeError):
            num_ok = False
            break

    if num_ok:
        precision = max(max_prec, max_scale + 1) + 2   # headroom
        return f"NUMERIC({precision},{max_scale})"

    # Date / timestamp
    if all(isinstance(v, (datetime, date)) for v in values):
        if all(isinstance(v, datetime) for v in values):
            return "TIMESTAMP"
        return "DATE"

    # Fallback: measure max length for VARCHAR
    max_len = max((len(str(v)) for v in values), default=0)
    if max_len <= 50:
        return "VARCHAR(50)"
    if max_len <= 255:
        return "VARCHAR(255)"
    return "TEXT"


def is_nullable(col_values: list, total_rows: int) -> bool:
    return len(col_values) < total_rows


def format_value(v, sql_type: str, dialect: str) -> str:
    """Render a Python cell value as a SQL literal."""
    if v is None:
        return "NULL"
    if sql_type in ("BOOLEAN", "TINYINT(1)"):
        if isinstance(v, bool):
            return "TRUE" if v else "FALSE"
        s = str(v).strip().lower()
        return "TRUE" if s in {"true", "yes", "1"} else "FALSE"
    if sql_type in ("TIMESTAMP", "DATE"):
        return f"'{v}'"
    if sql_type.startswith(("SMALLINT", "INTEGER", "BIGINT")):
        return str(int(float(str(v))))
    if sql_type.startswith("NUMERIC"):
        return str(Decimal(str(v)))
    # VARCHAR / TEXT — escape single quotes
    escaped = str(v).replace("'", "''")
    return f"'{escaped}'"


# ---------------------------------------------------------------------------
# Core generator
# ---------------------------------------------------------------------------

def sheet_to_sql(
    ws,
    table_name: str,
    dialect: str,
    header_row: int,
    no_drop: bool,
) -> str:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return f"-- Sheet '{ws.title}' is empty, skipped.\n"

    header_idx = header_row - 1
    headers = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(rows[header_idx])]
    col_names = [to_snake_case(h) for h in headers]

    data_rows = [r for r in rows[header_idx + 1:] if any(c is not None for c in r)]
    if not data_rows:
        return f"-- Sheet '{ws.title}': only a header row found, no data.\n"

    num_cols = len(col_names)

    # Collect non-null values per column for type inference
    col_values: list[list] = [[] for _ in range(num_cols)]
    for row in data_rows:
        for i in range(num_cols):
            v = row[i] if i < len(row) else None
            if v is not None and str(v).strip() != "":
                col_values[i].append(v)

    col_types = [
        infer_sql_type(col_values[i], dialect) if col_values[i] else "TEXT"
        for i in range(num_cols)
    ]
    nullability = [
        is_nullable(col_values[i], len(data_rows))
        for i in range(num_cols)
    ]

    lines = []
    lines.append(f"-- Auto-generated from sheet: {ws.title}")
    lines.append(f"-- Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"-- Rows: {len(data_rows)}, Columns: {num_cols}")
    lines.append("")

    if not no_drop:
        lines.append(f"DROP TABLE IF EXISTS {table_name};")
        lines.append("")

    lines.append(f"CREATE TABLE {table_name} (")
    col_defs = []
    max_name_len = max(len(n) for n in col_names)
    for i, (name, sql_type, nullable) in enumerate(zip(col_names, col_types, nullability)):
        padding = " " * (max_name_len - len(name) + 1)
        null_str = "" if nullable else " NOT NULL"
        col_defs.append(f"    {name}{padding}{sql_type}{null_str}")
    lines.append(",\n".join(col_defs))
    lines.append(");")
    lines.append("")

    # INSERT in batches of 500 rows
    batch_size = 500
    col_list = ", ".join(col_names)

    for batch_start in range(0, len(data_rows), batch_size):
        batch = data_rows[batch_start: batch_start + batch_size]
        lines.append(f"INSERT INTO {table_name} ({col_list})")
        lines.append("VALUES")
        value_rows = []
        for row in batch:
            vals = []
            for i in range(num_cols):
                v = row[i] if i < len(row) else None
                if v is not None and str(v).strip() == "":
                    v = None
                vals.append(format_value(v, col_types[i], dialect))
            value_rows.append("    (" + ", ".join(vals) + ")")
        lines.append(",\n".join(value_rows) + ";")
        lines.append("")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args():
    p = argparse.ArgumentParser(
        description="Generate SQL migration scripts from xlsx files.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    p.add_argument("xlsx", help="Path to the xlsx file")
    p.add_argument("--sheet", help="Sheet name or 0-based index (default: all sheets)")
    p.add_argument("--table", help="Override table name")
    p.add_argument("--dialect", choices=["pg", "mysql"], default="pg")
    p.add_argument("--out", help="Output directory (default: same dir as xlsx)")
    p.add_argument("--header", type=int, default=1, help="Header row number (default: 1)")
    p.add_argument("--no-drop", action="store_true", help="Skip DROP TABLE IF EXISTS")
    return p.parse_args()


def main():
    args = parse_args()
    xlsx_path = Path(args.xlsx).resolve()
    if not xlsx_path.exists():
        sys.exit(f"File not found: {xlsx_path}")

    out_dir = Path(args.out).resolve() if args.out else xlsx_path.parent
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"Loading {xlsx_path} ...")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # Determine which sheets to process
    if args.sheet is not None:
        try:
            idx = int(args.sheet)
            sheets = [wb.worksheets[idx]]
        except ValueError:
            if args.sheet not in wb.sheetnames:
                sys.exit(f"Sheet '{args.sheet}' not found. Available: {wb.sheetnames}")
            sheets = [wb[args.sheet]]
    else:
        sheets = wb.worksheets

    for ws in sheets:
        if args.table:
            table_name = args.table
        else:
            table_name = to_snake_case(ws.title)

        sql = sheet_to_sql(
            ws,
            table_name=table_name,
            dialect=args.dialect,
            header_row=args.header,
            no_drop=args.no_drop,
        )

        out_file = out_dir / f"{table_name}.sql"
        out_file.write_text(sql, encoding="utf-8")
        print(f"  -> {out_file}")

    print("Done.")


if __name__ == "__main__":
    main()
